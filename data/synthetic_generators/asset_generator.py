"""
AssetLiabilityGenerator: Generate Excel spreadsheets with wealth assessment data
Production-grade Excel with proper formatting
"""

import random
from typing import Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from .config import ASSET_PROPERTY_VALUES, VEHICLE_VALUES, LOAN_TYPES


class AssetLiabilityGenerator:
    """Generate realistic asset and liability Excel spreadsheets."""
    
    def __init__(self):
        """Initialize AssetLiabilityGenerator."""
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def generate_assets_liabilities(
        self,
        full_name: str,
        monthly_income: float,
        family_size: int,
        housing_type: str = "Rent",
        output_path: str = None
    ) -> Dict[str, Any]:
        """
        Generate assets and liabilities spreadsheet.
        
        Args:
            full_name: Person's full name
            monthly_income: Monthly income in AED
            family_size: Number of family members
            housing_type: Type of housing
            output_path: Path to save Excel file
            
        Returns:
            Dict: Asset/liability data and file path
        """
        # Generate assets
        assets = self._generate_assets(monthly_income, housing_type, family_size)
        
        # Generate liabilities
        liabilities = self._generate_liabilities(assets, monthly_income)
        
        # Calculate totals
        total_assets = sum(asset["value"] for asset in assets)
        total_liabilities = sum(liability["amount"] for liability in liabilities)
        net_worth = total_assets - total_liabilities
        
        data = {
            "full_name": full_name,
            "monthly_income": monthly_income,
            "family_size": family_size,
            "assets": assets,
            "liabilities": liabilities,
            "total_assets": total_assets,
            "total_liabilities": total_liabilities,
            "net_worth": net_worth,
            "asset_to_liability_ratio": total_assets / total_liabilities if total_liabilities > 0 else 0,
            "excel_path": output_path
        }
        
        # Generate Excel if path provided
        if output_path:
            self._generate_excel(data, output_path)
        
        return data
    
    def _generate_assets(
        self,
        monthly_income: float,
        housing_type: str,
        family_size: int
    ) -> list:
        """
        Generate realistic asset list.
        
        Args:
            monthly_income: Monthly income
            housing_type: Type of housing
            family_size: Family size
            
        Returns:
            List: Asset items with values
        """
        assets = []
        
        # Real Estate
        if housing_type == "Own (with mortgage)" or housing_type == "Own (paid off)":
            property_type = random.choice(list(ASSET_PROPERTY_VALUES.keys()))
            min_val, max_val = ASSET_PROPERTY_VALUES[property_type]
            property_value = round(random.uniform(min_val, max_val), 0)
            
            assets.append({
                "category": "Real Estate",
                "description": property_type,
                "value": property_value,
                "notes": "Primary residence"
            })
        
        # Vehicles
        num_vehicles = random.choices([0, 1, 2], weights=[0.3, 0.5, 0.2], k=1)[0]
        for i in range(num_vehicles):
            vehicle_type = random.choice(list(VEHICLE_VALUES.keys()))
            min_val, max_val = VEHICLE_VALUES[vehicle_type]
            vehicle_value = round(random.uniform(min_val, max_val), 0)
            
            assets.append({
                "category": "Vehicles",
                "description": vehicle_type,
                "value": vehicle_value,
                "notes": f"Vehicle {i+1}"
            })
        
        # Savings/Investments
        if monthly_income > 0:
            savings_ratio = random.uniform(0.5, 2.0)  # Months of salary in savings
            savings_amount = round(monthly_income * savings_ratio, 0)
            
            if savings_amount > 0:
                assets.append({
                    "category": "Liquid Assets",
                    "description": "Savings Account",
                    "value": savings_amount,
                    "notes": f"{savings_ratio:.1f} months salary"
                })
            
            # Investments (if income is higher)
            if monthly_income > 12000:
                investment_amount = round(random.uniform(50000, 200000), 0)
                assets.append({
                    "category": "Investments",
                    "description": "UAE Bonds/Stocks",
                    "value": investment_amount,
                    "notes": "Investment portfolio"
                })
        else:
            # Unemployed person with minimal savings
            if random.random() > 0.6:  # 40% chance of having small savings
                savings_amount = round(random.uniform(5000, 25000), 0)
                assets.append({
                    "category": "Liquid Assets",
                    "description": "Personal Savings",
                    "value": savings_amount,
                    "notes": "Emergency fund"
                })
        
        return assets
    
    def _generate_liabilities(self, assets: list, monthly_income: float) -> list:
        """
        Generate realistic liability list.
        
        Args:
            assets: List of assets
            monthly_income: Monthly income
            
        Returns:
            List: Liability items
        """
        liabilities = []
        
        # Mortgage (if property owner)
        property_assets = [a for a in assets if a["category"] == "Real Estate"]
        if property_assets:
            property_value = property_assets[0]["value"]
            # Assume 70% mortgage ratio
            mortgage_amount = round(property_value * 0.70, 0)
            remaining_years = random.randint(10, 25)
            monthly_payment = round(mortgage_amount / (remaining_years * 12), 0)
            
            liabilities.append({
                "category": "Mortgages",
                "description": "Home Mortgage",
                "amount": mortgage_amount,
                "monthly_payment": monthly_payment,
                "remaining_years": remaining_years
            })
        
        # Auto loans (if vehicles owned)
        vehicle_assets = [a for a in assets if a["category"] == "Vehicles"]
        for i, vehicle in enumerate(vehicle_assets):
            if random.random() > 0.3:  # 70% chance of having loan
                # Assume 60% auto loan ratio
                loan_amount = round(vehicle["value"] * 0.60, 0)
                remaining_years = random.randint(3, 7)
                monthly_payment = round(loan_amount / (remaining_years * 12), 0)
                
                liabilities.append({
                    "category": "Auto Loans",
                    "description": f"Vehicle Loan {i+1}",
                    "amount": loan_amount,
                    "monthly_payment": monthly_payment,
                    "remaining_years": remaining_years
                })
        
        # Personal loans
        if random.random() > 0.6:  # 40% chance
            loan_amount = round(random.uniform(20000, 100000), 0)
            remaining_years = random.randint(2, 5)
            monthly_payment = round(loan_amount / (remaining_years * 12), 0)
            
            liabilities.append({
                "category": "Personal Loans",
                "description": "Personal Loan",
                "amount": loan_amount,
                "monthly_payment": monthly_payment,
                "remaining_years": remaining_years
            })
        
        # Credit card debt
        cc_balance = round(random.uniform(5000, 50000), 0)
        liabilities.append({
            "category": "Credit Cards",
            "description": "Credit Card Balance",
            "amount": cc_balance,
            "monthly_payment": round(cc_balance * 0.05, 0),
            "remaining_years": 1  # Revolving
        })
        
        return liabilities
    
    def _generate_excel(self, data: Dict[str, Any], output_path: str) -> None:
        """
        Generate Excel spreadsheet.
        
        Args:
            data: Asset/liability data
            output_path: Path to save Excel
        """
        try:
            wb = Workbook()
            
            # Remove default sheet and create new ones
            wb.remove(wb.active)
            
            # Assets sheet
            ws_assets = wb.create_sheet("Assets")
            self._populate_assets_sheet(ws_assets, data)
            
            # Liabilities sheet
            ws_liabilities = wb.create_sheet("Liabilities")
            self._populate_liabilities_sheet(ws_liabilities, data)
            
            # Summary sheet
            ws_summary = wb.create_sheet("Summary", 0)
            self._populate_summary_sheet(ws_summary, data)
            
            # Save
            wb.save(output_path)
            
        except Exception as e:
            raise Exception(f"Error generating Excel: {str(e)}")
    
    def _populate_summary_sheet(self, ws, data: Dict[str, Any]) -> None:
        """Populate summary sheet."""
        # Header
        ws['A1'] = "FINANCIAL SUMMARY"
        ws['A1'].font = Font(bold=True, size=12, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="1f4788", end_color="1f4788", fill_type="solid")
        ws.merge_cells('A1:B1')
        
        # Details
        row = 3
        ws[f'A{row}'] = "Name:"
        ws[f'B{row}'] = data['full_name']
        row += 1
        
        ws[f'A{row}'] = "Monthly Income (AED):"
        ws[f'B{row}'] = data['monthly_income']
        ws[f'B{row}'].number_format = '#,##0'
        row += 2
        
        ws[f'A{row}'] = "Total Assets (AED):"
        ws[f'B{row}'] = data['total_assets']
        ws[f'B{row}'].number_format = '#,##0'
        ws[f'B{row}'].font = Font(bold=True)
        row += 1
        
        ws[f'A{row}'] = "Total Liabilities (AED):"
        ws[f'B{row}'] = data['total_liabilities']
        ws[f'B{row}'].number_format = '#,##0'
        ws[f'B{row}'].font = Font(bold=True)
        row += 1
        
        ws[f'A{row}'] = "Net Worth (AED):"
        ws[f'B{row}'] = data['net_worth']
        ws[f'B{row}'].number_format = '#,##0'
        ws[f'B{row}'].font = Font(bold=True, color="FFFFFF")
        ws[f'B{row}'].fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        row += 1
        
        ws[f'A{row}'] = "Asset/Liability Ratio:"
        ws[f'B{row}'] = round(data['asset_to_liability_ratio'], 2)
        ws[f'B{row}'].number_format = '0.00'
        
        # Set column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
    
    def _populate_assets_sheet(self, ws, data: Dict[str, Any]) -> None:
        """Populate assets sheet."""
        # Header
        headers = ["Category", "Description", "Value (AED)"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1f4788", end_color="1f4788", fill_type="solid")
            cell.border = self.thin_border
        
        # Data
        for row, asset in enumerate(data['assets'], 2):
            ws.cell(row=row, column=1).value = asset['category']
            ws.cell(row=row, column=2).value = asset['description']
            ws.cell(row=row, column=3).value = asset['value']
            ws.cell(row=row, column=3).number_format = '#,##0'
            
            for col in range(1, 4):
                ws.cell(row=row, column=col).border = self.thin_border
        
        # Total row
        total_row = len(data['assets']) + 2
        ws.cell(row=total_row, column=1).value = "TOTAL ASSETS"
        ws.cell(row=total_row, column=1).font = Font(bold=True)
        ws.cell(row=total_row, column=3).value = data['total_assets']
        ws.cell(row=total_row, column=3).number_format = '#,##0'
        ws.cell(row=total_row, column=3).font = Font(bold=True)
        ws.cell(row=total_row, column=3).fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        
        # Set column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 18
    
    def _populate_liabilities_sheet(self, ws, data: Dict[str, Any]) -> None:
        """Populate liabilities sheet."""
        # Header
        headers = ["Category", "Description", "Amount (AED)", "Monthly Payment", "Remaining Years"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="D32F2F", end_color="D32F2F", fill_type="solid")
            cell.border = self.thin_border
        
        # Data
        for row, liability in enumerate(data['liabilities'], 2):
            ws.cell(row=row, column=1).value = liability['category']
            ws.cell(row=row, column=2).value = liability['description']
            ws.cell(row=row, column=3).value = liability['amount']
            ws.cell(row=row, column=3).number_format = '#,##0'
            ws.cell(row=row, column=4).value = liability.get('monthly_payment', 0)
            ws.cell(row=row, column=4).number_format = '#,##0'
            ws.cell(row=row, column=5).value = liability.get('remaining_years', 0)
            
            for col in range(1, 6):
                ws.cell(row=row, column=col).border = self.thin_border
        
        # Total row
        total_row = len(data['liabilities']) + 2
        ws.cell(row=total_row, column=1).value = "TOTAL LIABILITIES"
        ws.cell(row=total_row, column=1).font = Font(bold=True)
        ws.cell(row=total_row, column=3).value = data['total_liabilities']
        ws.cell(row=total_row, column=3).number_format = '#,##0'
        ws.cell(row=total_row, column=3).font = Font(bold=True)
        ws.cell(row=total_row, column=3).fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
        
        # Set column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 16


# Quick test
if __name__ == "__main__":
    gen = AssetLiabilityGenerator()
    
    print("=" * 60)
    print("ASSET/LIABILITY GENERATION TEST")
    print("=" * 60)
    
    data = gen.generate_assets_liabilities(
        full_name="Ahmed Al Mazrouei",
        monthly_income=8500,
        family_size=4,
        housing_type="Own (with mortgage)"
    )
    
    print(f"\nAsset/Liability Data Generated:")
    print(f"  Total Assets: AED {data['total_assets']:,.0f}")
    print(f"  Total Liabilities: AED {data['total_liabilities']:,.0f}")
    print(f"  Net Worth: AED {data['net_worth']:,.0f}")
    print(f"  Ratio: {data['asset_to_liability_ratio']:.2f}")
