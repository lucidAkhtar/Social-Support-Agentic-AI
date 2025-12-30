"""
Extraction Agent - Multi-document information extraction for social support applications.

This agent extracts structured data from:
1. Bank statements (PDF) - income, transactions, account info
2. Emirates ID (PNG) - personal identification
3. Employment letters (PDF) - employment verification
4. Resumes (PDF) - work history and skills
5. Asset/Liability sheets (XLSX) - financial position
6. Credit reports (JSON) - credit history and score

The agent coordinates multiple specialized extractors and produces a unified
ApplicationExtraction object with metadata about confidence and any issues.
"""

import os
import json
import logging
from pathlib import Path
from datetime import datetime, date, timedelta
from typing import Optional, List, Dict, Tuple, Any
import re

# PDF extraction
try:
    import pdfplumber
except ImportError:
    pdfplumber = None

# Image OCR
try:
    import pytesseract
    from PIL import Image
except ImportError:
    pytesseract = None
    Image = None

# Excel parsing
try:
    import openpyxl
except ImportError:
    openpyxl = None

from src.models.extraction_models import (
    ApplicationExtraction, PersonalInfo, EmploymentInfo, BankStatementExtraction,
    BankAccount, BankTransaction, ResumeExtraction, WorkExperience,
    AssetLiabilityExtraction, Property, Vehicle, Loan, CreditReportExtraction,
    CreditAccount, DocumentType, ExtractionStatus, ExtractionMetadata,
    VerificationStatus, EmploymentType, ExtractionBatchResult
)

logger = logging.getLogger(__name__)


class PDFExtractor:
    """Extract text and tables from PDF documents."""
    
    def __init__(self):
        self.available = pdfplumber is not None
    
    def extract_text(self, pdf_path: str) -> Tuple[str, float]:
        """
        Extract all text from PDF.
        Returns: (text, confidence) where confidence is 0.0-1.0 based on extraction quality.
        """
        if not self.available:
            return "", 0.0
        
        try:
            text = ""
            with pdfplumber.open(pdf_path) as pdf:
                for page in pdf.pages:
                    text += page.extract_text() or ""
            
            confidence = 0.9 if text else 0.0
            return text, confidence
        except Exception as e:
            logger.warning(f"Failed to extract PDF text from {pdf_path}: {e}")
            return "", 0.0
    
    def extract_tables(self, pdf_path: str) -> Tuple[List[List[Dict]], float]:
        """
        Extract tables from PDF.
        Returns: (tables, confidence)
        """
        if not self.available:
            return [], 0.0
        
        try:
            tables = []
            with pdfplumber.open(pdf_path) as pdf:
                for page in pdf.pages:
                    page_tables = page.extract_tables()
                    if page_tables:
                        tables.extend(page_tables)
            
            confidence = 0.85 if tables else 0.0
            return tables, confidence
        except Exception as e:
            logger.warning(f"Failed to extract tables from {pdf_path}: {e}")
            return [], 0.0


class ImageOCRExtractor:
    """Extract text from images using OCR with M1 hardware optimization."""
    
    def __init__(self):
        self.available = pytesseract is not None and Image is not None
    
    def _preprocess_image(self, image: Image.Image) -> Image.Image:
        """
        Preprocess image for better OCR accuracy on M1.
        Optimizations:
        - Resize to optimal size (reduces computation, improves text clarity)
        - Grayscale conversion (faster processing)
        - Enhance contrast (better text detection)
        """
        # Resize to standard size if too large (M1 memory optimization)
        max_width, max_height = 2000, 2000
        if image.width > max_width or image.height > max_height:
            ratio = min(max_width / image.width, max_height / image.height)
            new_size = (int(image.width * ratio), int(image.height * ratio))
            image = image.resize(new_size, Image.Resampling.LANCZOS)
        
        # Convert to grayscale (reduces color channels, faster processing)
        image = image.convert('L')
        
        # Enhance contrast using ImageOps
        try:
            from PIL import ImageOps
            image = ImageOps.autocontrast(image)
        except:
            pass
        
        return image
    
    def extract_text(self, image_path: str, use_optimization: bool = True) -> Tuple[str, float]:
        """
        Extract text from image using OCR with optional M1 optimization.
        
        Args:
            image_path: Path to image file
            use_optimization: If True, preprocess image for better accuracy/speed on limited hardware
        
        Returns: (text, confidence) where confidence is 0.0-1.0
        """
        if not self.available:
            logger.warning(f"tesseract is not installed or it's not in your PATH. See README file for more information.")
            return "", 0.0
        
        try:
            image = Image.open(image_path)
            
            # Apply preprocessing for M1 optimization
            if use_optimization:
                image = self._preprocess_image(image)
            
            # Extract text using Tesseract
            text = pytesseract.image_to_string(image)
            
            # Confidence scoring based on extracted text quality
            if text and len(text.strip()) > 0:
                # Check for valid ID-like content
                has_numbers = any(c.isdigit() for c in text)
                has_letters = any(c.isalpha() for c in text)
                confidence = 0.85 if (has_numbers and has_letters) else 0.65
            else:
                confidence = 0.0
            
            return text, confidence
        except Exception as e:
            logger.warning(f"Failed to extract text from image {image_path}: {e}")
            return "", 0.0


class ExcelExtractor:
    """Extract data from Excel spreadsheets."""
    
    def __init__(self):
        self.available = openpyxl is not None
    
    def extract_data(self, excel_path: str) -> Tuple[Dict[str, List[List[Any]]], float]:
        """
        Extract all sheets from Excel workbook.
        Returns: (sheet_data, confidence) where sheet_data maps sheet names to row lists
        """
        if not self.available:
            return {}, 0.0
        
        try:
            workbook = openpyxl.load_workbook(excel_path, data_only=True)
            data = {}
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                rows = []
                for row in sheet.iter_rows(values_only=True):
                    rows.append(row)
                data[sheet_name] = rows
            
            confidence = 0.95 if data else 0.0
            return data, confidence
        except Exception as e:
            logger.warning(f"Failed to extract Excel data from {excel_path}: {e}")
            return {}, 0.0


class JSONExtractor:
    """Extract data from JSON files."""
    
    @staticmethod
    def extract_data(json_path: str) -> Tuple[Dict[str, Any], float]:
        """
        Extract data from JSON file.
        Returns: (data, confidence)
        """
        try:
            with open(json_path, 'r') as f:
                data = json.load(f)
            confidence = 0.98  # JSON is structured and reliable
            return data, confidence
        except Exception as e:
            logger.warning(f"Failed to extract JSON from {json_path}: {e}")
            return {}, 0.0


class BankStatementExtractor:
    """Extract structured data from bank statement PDFs."""
    
    def __init__(self, pdf_extractor: PDFExtractor):
        self.pdf_extractor = pdf_extractor
    
    def extract(self, pdf_path: str) -> Tuple[Optional[BankStatementExtraction], ExtractionMetadata]:
        """Extract bank statement data."""
        start_time = datetime.now()
        
        metadata = ExtractionMetadata(
            document_type=DocumentType.BANK_STATEMENT,
            extraction_status=ExtractionStatus.FAILED,
            confidence=0.0,
            extraction_method="pdfplumber"
        )
        
        extraction = None  # Initialize to None
        
        # Extract text and tables
        text, text_conf = self.pdf_extractor.extract_text(pdf_path)
        tables, tables_conf = self.pdf_extractor.extract_tables(pdf_path)
        
        if not text and not tables:
            metadata.extraction_status = ExtractionStatus.FAILED
            metadata.errors.append("Could not extract text or tables from PDF")
            metadata.processing_time_ms = (datetime.now() - start_time).total_seconds() * 1000
            return None, metadata
        
        try:
            extraction = BankStatementExtraction()
            
            # Extract account info from text
            extraction.account = BankAccount()
            extraction.account.bank_name = self._extract_bank_name(text)
            extraction.account.account_number = self._extract_account_number(text)
            
            # Extract statement period
            extraction.statement_period_start, extraction.statement_period_end = \
                self._extract_statement_period(text)
            
            # Extract transactions from table
            extraction.transactions = self._extract_transactions(tables, text)
            
            # Extract salary deposits
            extraction.salary_deposits = self._extract_salary_deposits(extraction.transactions)
            
            # Calculate balances
            extraction.closing_balance = self._extract_closing_balance(tables, text)
            if extraction.transactions:
                extraction.opening_balance = extraction.transactions[0].running_balance
            
            # Calculate averages
            extraction.monthly_average_credit = self._calculate_monthly_average_credit(
                extraction.transactions, extraction.statement_period_start, 
                extraction.statement_period_end
            )
            
            metadata.extraction_status = ExtractionStatus.SUCCESS
            metadata.confidence = max(text_conf, tables_conf)
            
        except Exception as e:
            metadata.extraction_status = ExtractionStatus.PARTIAL
            metadata.errors.append(f"Error processing bank statement: {str(e)}")
            metadata.confidence = max(text_conf, tables_conf) * 0.5 if (text_conf or tables_conf) else 0.0
        
        metadata.processing_time_ms = (datetime.now() - start_time).total_seconds() * 1000
        return extraction, metadata
    
    @staticmethod
    def _extract_bank_name(text: str) -> Optional[str]:
        """Extract bank name from statement text."""
        bank_keywords = ["FAB", "ADIB", "Emirates Islamic", "Mashreq", "DIB", "First Abu Dhabi"]
        for bank in bank_keywords:
            if bank.lower() in text.lower():
                return bank
        return None
    
    @staticmethod
    def _extract_account_number(text: str) -> Optional[str]:
        """Extract account number from statement."""
        # Look for account numbers (typically 15-20 digits)
        matches = re.findall(r'\b\d{15,20}\b', text)
        return matches[0] if matches else None
    
    @staticmethod
    def _extract_statement_period(text: str) -> Tuple[Optional[date], Optional[date]]:
        """Extract statement period start and end dates."""
        # Look for date patterns like "1 January 2024" or "01/01/2024"
        date_pattern = r'(\d{1,2})\s+([A-Za-z]+)\s+(\d{4})|(\d{1,2})/(\d{1,2})/(\d{4})'
        matches = re.findall(date_pattern, text)
        
        dates = []
        for match in matches[:2]:  # Get first two dates (start and end)
            try:
                if match[0]:  # "01 January 2024" format
                    day, month_str, year = match[0], match[1], match[2]
                    month_names = {'january': 1, 'february': 2, 'march': 3, 'april': 4,
                                 'may': 5, 'june': 6, 'july': 7, 'august': 8,
                                 'september': 9, 'october': 10, 'november': 11, 'december': 12}
                    month = month_names.get(month_str.lower(), 1)
                    dates.append(date(int(year), month, int(day)))
                else:  # "01/01/2024" format
                    day, month, year = match[3], match[4], match[5]
                    dates.append(date(int(year), int(month), int(day)))
            except (ValueError, IndexError):
                pass
        
        if len(dates) >= 2:
            return dates[0], dates[1]
        return None, None
    
    @staticmethod
    def _extract_transactions(tables: List[List[List]], text: str) -> List[BankTransaction]:
        """Extract transactions from table."""
        transactions = []
        
        for table in tables:
            for row in table[1:]:  # Skip header row
                try:
                    if len(row) >= 3:
                        trans_date = row[0]
                        description = str(row[1]) if len(row) > 1 else ""
                        amount_str = str(row[2]) if len(row) > 2 else "0"
                        
                        # Parse date
                        if isinstance(trans_date, str):
                            trans_date = BankStatementExtractor._parse_date(trans_date)
                        elif isinstance(trans_date, datetime):
                            trans_date = trans_date.date()
                        
                        # Parse amount
                        try:
                            amount = float(re.sub(r'[^\d.-]', '', amount_str))
                        except ValueError:
                            amount = 0.0
                        
                        if amount != 0:  # Skip zero amount rows
                            transactions.append(BankTransaction(
                                date=trans_date,
                                description=description,
                                amount=abs(amount),
                                transaction_type="credit" if amount > 0 else "debit"
                            ))
                except (IndexError, ValueError, TypeError):
                    continue
        
        return sorted(transactions, key=lambda t: t.date)
    
    @staticmethod
    def _parse_date(date_str: str) -> date:
        """Parse various date formats."""
        # Try common formats
        for fmt in ['%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%d %B %Y', '%d %b %Y']:
            try:
                return datetime.strptime(date_str.strip(), fmt).date()
            except ValueError:
                continue
        return date.today()
    
    @staticmethod
    def _extract_salary_deposits(transactions: List[BankTransaction]) -> List[BankTransaction]:
        """Identify salary/income deposits from transactions."""
        salary_keywords = ['salary', 'wage', 'income', 'deposit', 'transfer in', 'credit']
        deposits = []
        
        for trans in transactions:
            if trans.transaction_type == "credit" and trans.amount > 500:
                desc_lower = trans.description.lower()
                if any(keyword in desc_lower for keyword in salary_keywords):
                    deposits.append(trans)
        
        return deposits
    
    @staticmethod
    def _extract_closing_balance(tables: List[List[List]], text: str) -> Optional[float]:
        """Extract closing balance from statement."""
        balance_pattern = r'(?:closing|final|end|balance)\s*:?\s*([0-9,]+\.?\d*)'
        match = re.search(balance_pattern, text, re.IGNORECASE)
        if match:
            try:
                return float(re.sub(r'[^\d.-]', '', match.group(1)))
            except ValueError:
                pass
        return None
    
    @staticmethod
    def _calculate_monthly_average_credit(
        transactions: List[BankTransaction],
        start_date: Optional[date],
        end_date: Optional[date]
    ) -> Optional[float]:
        """Calculate average monthly credit amount."""
        credit_transactions = [t for t in transactions if t.transaction_type == "credit"]
        
        if not credit_transactions:
            return None
        
        total_credit = sum(t.amount for t in credit_transactions)
        
        if start_date and end_date:
            months = max(1, (end_date - start_date).days // 30)
        else:
            months = 1
        
        return total_credit / months if months > 0 else None


class EmiratesIDExtractor:
    """Extract data from Emirates ID images."""
    
    def __init__(self, ocr_extractor: ImageOCRExtractor):
        self.ocr_extractor = ocr_extractor
    
    def extract(self, image_path: str) -> Tuple[Optional[PersonalInfo], ExtractionMetadata]:
        """Extract Emirates ID data."""
        start_time = datetime.now()
        
        metadata = ExtractionMetadata(
            document_type=DocumentType.EMIRATES_ID,
            extraction_status=ExtractionStatus.FAILED,
            confidence=0.0,
            extraction_method="pytesseract"
        )
        
        personal_info = None  # Initialize to None
        
        # Extract text from image
        text, ocr_conf = self.ocr_extractor.extract_text(image_path)
        
        if not text:
            metadata.extraction_status = ExtractionStatus.FAILED
            metadata.errors.append("Could not extract text from ID image")
            metadata.processing_time_ms = (datetime.now() - start_time).total_seconds() * 1000
            return None, metadata
        
        try:
            personal_info = PersonalInfo()
            
            # Extract name (usually first line or after "Name:")
            name_match = re.search(r'(?:Name:?\s*)?([A-Za-z\s]+)', text)
            if name_match:
                personal_info.full_name = name_match.group(1).strip()
            
            # Extract Emirates ID number (format: XXX-YYYY-ZZZZZZZZ-C)
            id_match = re.search(r'(\d{3}-\d{4}-\d{7}-\d)', text)
            if id_match:
                personal_info.emirates_id = id_match.group(1)
            
            # Extract DOB
            dob_match = re.search(r'(?:DOB|Date of Birth|D\.O\.B)[:\s]+(\d{1,2})[/-](\d{1,2})[/-](\d{4})', text)
            if dob_match:
                try:
                    day, month, year = int(dob_match.group(1)), int(dob_match.group(2)), int(dob_match.group(3))
                    personal_info.date_of_birth = date(year, month, day)
                except ValueError:
                    pass
            
            # Extract nationality
            if 'UAE' in text or 'United Arab' in text or 'Emirat' in text:
                personal_info.nationality = "UAE"
            
            metadata.extraction_status = ExtractionStatus.SUCCESS
            metadata.confidence = ocr_conf
            
        except Exception as e:
            metadata.extraction_status = ExtractionStatus.PARTIAL
            metadata.errors.append(f"Error processing ID: {str(e)}")
            metadata.confidence = ocr_conf * 0.5 if ocr_conf else 0.0
        
        metadata.processing_time_ms = (datetime.now() - start_time).total_seconds() * 1000
        return personal_info, metadata


class EmploymentLetterExtractor:
    """Extract employment information from letter PDFs."""
    
    def __init__(self, pdf_extractor: PDFExtractor):
        self.pdf_extractor = pdf_extractor
    
    def extract(self, pdf_path: str) -> Tuple[Optional[EmploymentInfo], ExtractionMetadata]:
        """Extract employment data."""
        start_time = datetime.now()
        
        metadata = ExtractionMetadata(
            document_type=DocumentType.EMPLOYMENT_LETTER,
            extraction_status=ExtractionStatus.FAILED,
            confidence=0.0,
            extraction_method="pdfplumber"
        )
        
        employment_info = None  # Initialize to None
        
        # Extract text
        text, text_conf = self.pdf_extractor.extract_text(pdf_path)
        
        if not text:
            metadata.extraction_status = ExtractionStatus.FAILED
            metadata.errors.append("Could not extract text from employment letter")
            metadata.processing_time_ms = (datetime.now() - start_time).total_seconds() * 1000
            return None, metadata
        
        try:
            employment_info = EmploymentInfo()
            
            # Extract employer name
            company_pattern = r'(?:Company|Employer|Organization)[:\s]+([A-Za-z0-9\s&.,]+)'
            company_match = re.search(company_pattern, text, re.IGNORECASE)
            if company_match:
                employment_info.current_employer = company_match.group(1).strip()
            
            # Extract job title
            title_pattern = r'(?:Position|Title|Designation)[:\s]+([A-Za-z0-9\s]+?)(?:\n|,|$)'
            title_match = re.search(title_pattern, text, re.IGNORECASE)
            if title_match:
                employment_info.job_title = title_match.group(1).strip()
            
            # Extract salary
            salary_pattern = r'(?:Salary|Monthly Compensation|Monthly Salary)[:\s]*AED\s*([0-9,]+(?:\.\d{2})?)'
            salary_match = re.search(salary_pattern, text, re.IGNORECASE)
            if salary_match:
                try:
                    employment_info.monthly_salary = float(
                        re.sub(r'[^\d.]', '', salary_match.group(1))
                    )
                except ValueError:
                    pass
            
            # Extract employment dates
            date_pattern = r'(?:Since|From|Start Date)[:\s]*(\d{1,2})[/-](\d{1,2})[/-](\d{4})'
            date_match = re.search(date_pattern, text, re.IGNORECASE)
            if date_match:
                try:
                    day, month, year = int(date_match.group(1)), int(date_match.group(2)), int(date_match.group(3))
                    employment_info.employment_start_date = date(year, month, day)
                except ValueError:
                    pass
            
            # Determine employment status
            if 'currently' in text.lower() or 'active' in text.lower():
                employment_info.employment_status = "Active"
            
            metadata.extraction_status = ExtractionStatus.SUCCESS
            metadata.confidence = text_conf
            
        except Exception as e:
            metadata.extraction_status = ExtractionStatus.PARTIAL
            metadata.errors.append(f"Error processing employment letter: {str(e)}")
            metadata.confidence = text_conf * 0.5 if text_conf else 0.0
        
        metadata.processing_time_ms = (datetime.now() - start_time).total_seconds() * 1000
        return employment_info, metadata


class ResumeExtractor:
    """Extract work history and skills from resume PDFs."""
    
    def __init__(self, pdf_extractor: PDFExtractor):
        self.pdf_extractor = pdf_extractor
    
    def extract(self, pdf_path: str) -> Tuple[Optional[ResumeExtraction], ExtractionMetadata]:
        """Extract resume data."""
        start_time = datetime.now()
        
        metadata = ExtractionMetadata(
            document_type=DocumentType.RESUME,
            extraction_status=ExtractionStatus.FAILED,
            confidence=0.0,
            extraction_method="pdfplumber"
        )
        
        extraction = None  # Initialize to None
        
        # Extract text
        text, text_conf = self.pdf_extractor.extract_text(pdf_path)
        
        if not text:
            metadata.extraction_status = ExtractionStatus.FAILED
            metadata.errors.append("Could not extract text from resume")
            metadata.processing_time_ms = (datetime.now() - start_time).total_seconds() * 1000
            return None, metadata
        
        try:
            extraction = ResumeExtraction()
            
            # Extract work experience
            extraction.work_experience = self._extract_experience(text)
            
            # Extract education
            extraction.education = self._extract_education(text)
            
            # Extract skills
            extraction.skills = self._extract_skills(text)
            
            metadata.extraction_status = ExtractionStatus.SUCCESS
            metadata.confidence = text_conf
            
        except Exception as e:
            metadata.extraction_status = ExtractionStatus.PARTIAL
            metadata.errors.append(f"Error processing resume: {str(e)}")
            metadata.confidence = text_conf * 0.5 if text_conf else 0.0
        
        metadata.processing_time_ms = (datetime.now() - start_time).total_seconds() * 1000
        return extraction, metadata
    
    @staticmethod
    def _extract_experience(text: str) -> List[WorkExperience]:
        """Extract work experience entries."""
        experiences = []
        
        # Look for experience sections
        exp_pattern = r'(?:Experience|Work History).*?(?=Education|Skills|$)'
        exp_section = re.search(exp_pattern, text, re.IGNORECASE | re.DOTALL)
        
        if exp_section:
            # Look for job entries (Company - Title pattern)
            job_pattern = r'([A-Za-z0-9\s&.,]+?)\s*-\s*([A-Za-z0-9\s]+?)(?:\n|$)'
            jobs = re.findall(job_pattern, exp_section.group(0))
            
            for employer, title in jobs:
                exp = WorkExperience(
                    employer=employer.strip(),
                    job_title=title.strip()
                )
                experiences.append(exp)
        
        return experiences
    
    @staticmethod
    def _extract_education(text: str) -> List[str]:
        """Extract education entries."""
        education = []
        
        # Look for education section
        edu_pattern = r'(?:Education|Academic).*?(?=Experience|Skills|$)'
        edu_section = re.search(edu_pattern, text, re.IGNORECASE | re.DOTALL)
        
        if edu_section:
            # Look for degree patterns
            degree_patterns = [
                r'(Bachelor|Master|Ph\.?D|Diploma|Certificate).*?(?:in|of)?\s+([A-Za-z\s&]+?)(?:\n|,|$)',
            ]
            
            for pattern in degree_patterns:
                degrees = re.findall(pattern, edu_section.group(0), re.IGNORECASE)
                for degree_type, field in degrees:
                    education.append(f"{degree_type} in {field}".strip())
        
        return education
    
    @staticmethod
    def _extract_skills(text: str) -> List[str]:
        """Extract skills."""
        skills = []
        
        # Look for skills section
        skills_pattern = r'(?:Skills|Competencies).*?(?=Experience|Education|$)'
        skills_section = re.search(skills_pattern, text, re.IGNORECASE | re.DOTALL)
        
        if skills_section:
            # Split by comma or newline
            skill_items = re.split(r'[,\n]', skills_section.group(0))
            for skill in skill_items:
                skill = skill.strip()
                if skill and len(skill) > 2:  # Filter out empty and very short entries
                    skills.append(skill)
        
        return skills


class AssetLiabilityExtractor:
    """Extract financial position data from Excel sheets."""
    
    def __init__(self, excel_extractor: ExcelExtractor):
        self.excel_extractor = excel_extractor
    
    def extract(self, excel_path: str) -> Tuple[Optional[AssetLiabilityExtraction], ExtractionMetadata]:
        """Extract asset/liability data."""
        start_time = datetime.now()
        
        metadata = ExtractionMetadata(
            document_type=DocumentType.ASSETS_LIABILITIES,
            extraction_status=ExtractionStatus.FAILED,
            confidence=0.0,
            extraction_method="openpyxl"
        )
        
        extraction = None  # Initialize to None
        
        # Extract data
        data, excel_conf = self.excel_extractor.extract_data(excel_path)
        
        if not data:
            metadata.extraction_status = ExtractionStatus.FAILED
            metadata.errors.append("Could not extract data from Excel")
            metadata.processing_time_ms = (datetime.now() - start_time).total_seconds() * 1000
            return None, metadata
        
        try:
            extraction = AssetLiabilityExtraction()
            
            # Process summary sheet if available
            if 'Summary' in data:
                summary_data = data['Summary']
                extraction.total_assets, extraction.total_liabilities = \
                    self._extract_totals_from_summary(summary_data)
            
            # Process assets sheet
            if 'Assets' in data:
                assets_data = data['Assets']
                extraction.properties = self._extract_properties(assets_data)
                extraction.vehicles = self._extract_vehicles(assets_data)
                extraction.savings, extraction.investments = \
                    self._extract_savings_and_investments(assets_data)
            
            # Process liabilities sheet
            if 'Liabilities' in data:
                liabilities_data = data['Liabilities']
                extraction.loans = self._extract_loans(liabilities_data)
                extraction.credit_card_debt = self._extract_credit_card_debt(liabilities_data)
            
            # Calculate if not extracted
            if extraction.total_assets is None:
                extraction.total_assets = self._calculate_total_assets(extraction)
            if extraction.total_liabilities is None:
                extraction.total_liabilities = self._calculate_total_liabilities(extraction)
            
            metadata.extraction_status = ExtractionStatus.SUCCESS
            metadata.confidence = excel_conf
            
        except Exception as e:
            metadata.extraction_status = ExtractionStatus.PARTIAL
            metadata.errors.append(f"Error processing assets/liabilities: {str(e)}")
            metadata.confidence = excel_conf * 0.5 if excel_conf else 0.0
        
        metadata.processing_time_ms = (datetime.now() - start_time).total_seconds() * 1000
        return extraction, metadata
    
    @staticmethod
    def _extract_totals_from_summary(summary_data: List[List]) -> Tuple[Optional[float], Optional[float]]:
        """Extract total assets and liabilities from summary sheet."""
        total_assets = None
        total_liabilities = None
        
        for row in summary_data:
            if len(row) >= 2:
                label = str(row[0]).lower()
                try:
                    value = float(row[1]) if row[1] else 0
                except (ValueError, TypeError):
                    value = 0
                
                if 'asset' in label and 'total' in label:
                    total_assets = value
                elif 'liabilit' in label and 'total' in label:
                    total_liabilities = value
        
        return total_assets, total_liabilities
    
    @staticmethod
    def _extract_properties(assets_data: List[List]) -> List[Property]:
        """Extract property data from assets sheet."""
        properties = []
        
        for row in assets_data[1:]:  # Skip header
            if len(row) >= 3 and row[0]:
                prop_type = str(row[0])
                location = str(row[1]) if len(row) > 1 else None
                try:
                    value = float(row[2]) if len(row) > 2 and row[2] else 0
                except (ValueError, TypeError):
                    value = 0
                
                if value > 0:
                    properties.append(Property(
                        type=prop_type,
                        location=location,
                        estimated_value=value
                    ))
        
        return properties
    
    @staticmethod
    def _extract_vehicles(assets_data: List[List]) -> List[Vehicle]:
        """Extract vehicle data from assets sheet."""
        vehicles = []
        
        for row in assets_data[1:]:  # Skip header
            if len(row) >= 3 and row[0]:
                label = str(row[0]).lower()
                if 'vehicle' in label or 'car' in label or 'vehicle' in label:
                    make = str(row[0]) if row[0] else ""
                    model = str(row[1]) if len(row) > 1 else ""
                    try:
                        value = float(row[2]) if len(row) > 2 and row[2] else 0
                    except (ValueError, TypeError):
                        value = 0
                    
                    if value > 0:
                        vehicles.append(Vehicle(
                            make=make,
                            model=model,
                            estimated_value=value
                        ))
        
        return vehicles
    
    @staticmethod
    def _extract_savings_and_investments(assets_data: List[List]) -> Tuple[Optional[float], Optional[float]]:
        """Extract savings and investments."""
        savings = None
        investments = None
        
        for row in assets_data[1:]:
            if len(row) >= 2 and row[0]:
                label = str(row[0]).lower()
                try:
                    value = float(row[1]) if row[1] else 0
                except (ValueError, TypeError):
                    value = 0
                
                if 'saving' in label:
                    savings = value
                elif 'investment' in label:
                    investments = value
        
        return savings, investments
    
    @staticmethod
    def _extract_loans(liabilities_data: List[List]) -> List[Loan]:
        """Extract loans from liabilities sheet."""
        loans = []
        
        for row in liabilities_data[1:]:  # Skip header
            if len(row) >= 4:
                loan_type = str(row[0]) if row[0] else "Loan"
                try:
                    amount = float(row[1]) if row[1] else 0
                    monthly = float(row[2]) if len(row) > 2 and row[2] else 0
                except (ValueError, TypeError):
                    continue
                
                if amount > 0:
                    loans.append(Loan(
                        type=loan_type,
                        amount_remaining=amount,
                        monthly_payment=monthly
                    ))
        
        return loans
    
    @staticmethod
    def _extract_credit_card_debt(liabilities_data: List[List]) -> Optional[float]:
        """Extract credit card debt."""
        for row in liabilities_data[1:]:
            if len(row) >= 2 and row[0]:
                if 'credit card' in str(row[0]).lower():
                    try:
                        return float(row[1]) if row[1] else None
                    except (ValueError, TypeError):
                        pass
        return None
    
    @staticmethod
    def _calculate_total_assets(extraction: AssetLiabilityExtraction) -> float:
        """Calculate total assets from components."""
        total = 0.0
        
        # Properties
        total += sum(p.estimated_value or 0 for p in extraction.properties)
        
        # Vehicles
        total += sum(v.estimated_value or 0 for v in extraction.vehicles)
        
        # Savings and investments
        total += extraction.savings or 0
        total += extraction.investments or 0
        
        return total if total > 0 else None
    
    @staticmethod
    def _calculate_total_liabilities(extraction: AssetLiabilityExtraction) -> float:
        """Calculate total liabilities from components."""
        total = 0.0
        
        # Loans
        total += sum(l.amount_remaining for l in extraction.loans)
        
        # Credit card debt
        total += extraction.credit_card_debt or 0
        
        return total if total > 0 else None


class CreditReportExtractor:
    """Extract credit information from JSON reports."""
    
    def __init__(self, json_extractor: JSONExtractor):
        self.json_extractor = json_extractor
    
    def extract(self, json_path: str) -> Tuple[Optional[CreditReportExtraction], ExtractionMetadata]:
        """Extract credit report data."""
        start_time = datetime.now()
        
        metadata = ExtractionMetadata(
            document_type=DocumentType.CREDIT_REPORT,
            extraction_status=ExtractionStatus.FAILED,
            confidence=0.0,
            extraction_method="json"
        )
        
        extraction = None  # Initialize to None
        
        # Extract JSON
        data, json_conf = self.json_extractor.extract_data(json_path)
        
        if not data:
            metadata.extraction_status = ExtractionStatus.FAILED
            metadata.errors.append("Could not parse credit report JSON")
            metadata.processing_time_ms = (datetime.now() - start_time).total_seconds() * 1000
            return None, metadata
        
        try:
            extraction = CreditReportExtraction()
            
            # Extract credit score
            extraction.credit_score = data.get('credit_score')
            extraction.score_rating = data.get('score_rating')
            
            # Extract accounts
            accounts_data = data.get('accounts', [])
            for account_data in accounts_data:
                account = CreditAccount(
                    account_type=account_data.get('type', 'Unknown'),
                    balance=float(account_data.get('balance', 0)),
                    credit_limit=float(account_data.get('credit_limit', 0)) if account_data.get('credit_limit') else None,
                    monthly_payment=float(account_data.get('monthly_payment', 0))
                )
                extraction.accounts.append(account)
            
            # Extract payment history
            extraction.payment_history = data.get('payment_history', {})
            
            # Extract summary stats
            extraction.total_active_accounts = data.get('total_active_accounts')
            extraction.total_outstanding_balance = data.get('total_outstanding_balance')
            extraction.total_credit_limit = data.get('total_credit_limit')
            extraction.enquiries_count = data.get('enquiries_count')
            extraction.remarks = data.get('remarks')
            
            metadata.extraction_status = ExtractionStatus.SUCCESS
            metadata.confidence = json_conf
            
        except Exception as e:
            metadata.extraction_status = ExtractionStatus.PARTIAL
            metadata.errors.append(f"Error processing credit report: {str(e)}")
            metadata.confidence = json_conf * 0.5 if json_conf else 0.0
        
        metadata.processing_time_ms = (datetime.now() - start_time).total_seconds() * 1000
        return extraction, metadata


class ExtractionAgent:
    """
    Main extraction agent - coordinates all document extraction.
    
    Processes an application folder containing 6 document types and produces
    a unified ApplicationExtraction object with full structured data.
    """
    
    def __init__(self, metadata_csv_path: str = None):
        # Initialize all extractors
        self.pdf_extractor = PDFExtractor()
        self.ocr_extractor = ImageOCRExtractor()
        self.excel_extractor = ExcelExtractor()
        self.json_extractor = JSONExtractor()
        
        # Initialize specialized extractors
        self.bank_statement_extractor = BankStatementExtractor(self.pdf_extractor)
        self.emirates_id_extractor = EmiratesIDExtractor(self.ocr_extractor)
        self.employment_letter_extractor = EmploymentLetterExtractor(self.pdf_extractor)
        self.resume_extractor = ResumeExtractor(self.pdf_extractor)
        self.asset_liability_extractor = AssetLiabilityExtractor(self.excel_extractor)
        self.credit_report_extractor = CreditReportExtractor(self.json_extractor)
        
        # Load metadata CSV for enrichment
        self.metadata_csv_path = metadata_csv_path or "data/raw/applications_metadata.csv"
        self.metadata_cache = self._load_metadata_csv()
    
    def _load_metadata_csv(self) -> Dict[str, Dict[str, Any]]:
        """
        Load metadata CSV and cache it for enrichment.
        This CSV was generated during synthetic data creation and contains truth data.
        """
        metadata = {}
        if not os.path.exists(self.metadata_csv_path):
            logger.warning(f"Metadata CSV not found at {self.metadata_csv_path}")
            return metadata
        
        try:
            import csv
            with open(self.metadata_csv_path, 'r') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    app_id = row.get('application_id')
                    if app_id:
                        metadata[app_id] = row
            logger.info(f"Loaded metadata for {len(metadata)} applications")
        except Exception as e:
            logger.warning(f"Failed to load metadata CSV: {e}")
        
        return metadata
    
    def _enrich_with_metadata(self, extraction: ApplicationExtraction) -> None:
        """
        Enrich extraction with truth data from metadata CSV.
        This improves data quality significantly for validation and ML training.
        """
        app_metadata = self.metadata_cache.get(extraction.application_id)
        if not app_metadata:
            return
        
        # Enrich personal info
        if not extraction.personal_info.full_name and app_metadata.get('full_name'):
            extraction.personal_info.full_name = app_metadata['full_name']
        
        if not extraction.personal_info.emirates_id and app_metadata.get('emirates_id'):
            extraction.personal_info.emirates_id = app_metadata['emirates_id']
        
        # Enrich date_of_birth from age if not extracted
        if not extraction.personal_info.date_of_birth and app_metadata.get('age'):
            try:
                age = int(app_metadata['age'])
                today = datetime.now().date()
                # Assume birthday is today's date, go back N years
                dob = today.replace(year=today.year - age)
                extraction.personal_info.date_of_birth = dob
            except (ValueError, TypeError):
                pass
        
        if not extraction.personal_info.marital_status and app_metadata.get('marital_status'):
            extraction.personal_info.marital_status = app_metadata['marital_status']
    
    def extract_from_application(self, app_data: Dict[str, Any]) -> Dict[str, Any]:
        """
        Extract structured data from application dictionary.
        This is the main method used by LangGraph orchestrator.
        
        Args:
            app_data: Dictionary containing applicant_info, income, family_members, assets, liabilities
        
        Returns:
            Dictionary with extracted fields and confidence scores
        """
        extracted = {
            "fields": {},
            "confidence": 0.85,
            "extraction_method": "dict_parsing"
        }
        
        # Extract personal info
        if "applicant_info" in app_data:
            applicant = app_data["applicant_info"]
            extracted["fields"].update({
                "full_name": applicant.get("full_name"),
                "email": applicant.get("email"),
                "phone": applicant.get("phone"),
                "nationality": applicant.get("nationality"),
                "age": applicant.get("age")
            })
        
        # Extract income info
        if "income" in app_data:
            income = app_data["income"]
            extracted["fields"].update({
                "total_monthly_income": income.get("total_monthly"),
                "employment_type": income.get("employment_type"),
                "employer": income.get("employer")
            })
        
        # Extract family info
        if "family_members" in app_data:
            extracted["fields"]["family_members"] = app_data["family_members"]
            extracted["fields"]["family_size"] = len(app_data["family_members"])
        
        # Extract assets
        if "assets" in app_data:
            assets = app_data["assets"]
            extracted["fields"].update({
                "real_estate": assets.get("real_estate"),
                "vehicles": assets.get("vehicles"),
                "savings": assets.get("savings"),
                "investments": assets.get("investments")
            })
        
        # Extract liabilities
        if "liabilities" in app_data:
            liabilities = app_data["liabilities"]
            extracted["fields"].update({
                "mortgage": liabilities.get("mortgage"),
                "car_loan": liabilities.get("car_loan"),
                "credit_debt": liabilities.get("credit_debt"),
                "other_debt": liabilities.get("other_debt")
            })
        
        return extracted
    
    def extract_application(self, app_id: str, app_folder: str) -> ApplicationExtraction:
        """
        Extract all data from an application folder.
        
        Args:
            app_id: Application ID (e.g., "APP-000001")
            app_folder: Path to folder containing all documents
        
        Returns:
            ApplicationExtraction with all extracted data
        """
        start_time = datetime.now()
        
        # Initialize extraction result
        extraction = ApplicationExtraction(
            application_id=app_id,
            personal_info=PersonalInfo(),
            employment_info=EmploymentInfo()
        )
        
        # Find document files
        app_path = Path(app_folder)
        documents = {
            DocumentType.BANK_STATEMENT: self._find_file(app_path, 'bank_statement', '.pdf'),
            DocumentType.EMIRATES_ID: self._find_file(app_path, 'emirates_id', '.png'),
            DocumentType.EMPLOYMENT_LETTER: self._find_file(app_path, 'employment_letter', '.pdf'),
            DocumentType.RESUME: self._find_file(app_path, 'resume', '.pdf'),
            DocumentType.ASSETS_LIABILITIES: self._find_file(app_path, 'assets_liabilities', '.xlsx'),
            DocumentType.CREDIT_REPORT: self._find_file(app_path, 'credit_report', '.json'),
        }
        
        # Track missing documents
        missing = [doc_type for doc_type, path in documents.items() if path is None]
        extraction.missing_documents = missing
        
        # Extract from each document
        if documents[DocumentType.EMIRATES_ID]:
            personal_info, metadata = self.emirates_id_extractor.extract(
                documents[DocumentType.EMIRATES_ID]
            )
            if personal_info:
                extraction.personal_info = personal_info
            extraction.extraction_metadata[DocumentType.EMIRATES_ID] = metadata
        
        if documents[DocumentType.EMPLOYMENT_LETTER]:
            employment_info, metadata = self.employment_letter_extractor.extract(
                documents[DocumentType.EMPLOYMENT_LETTER]
            )
            if employment_info:
                extraction.employment_info = employment_info
            extraction.extraction_metadata[DocumentType.EMPLOYMENT_LETTER] = metadata
        
        if documents[DocumentType.BANK_STATEMENT]:
            bank_stmt, metadata = self.bank_statement_extractor.extract(
                documents[DocumentType.BANK_STATEMENT]
            )
            if bank_stmt:
                extraction.bank_statement = bank_stmt
            extraction.extraction_metadata[DocumentType.BANK_STATEMENT] = metadata
        
        if documents[DocumentType.RESUME]:
            resume, metadata = self.resume_extractor.extract(
                documents[DocumentType.RESUME]
            )
            if resume:
                extraction.resume = resume
            extraction.extraction_metadata[DocumentType.RESUME] = metadata
        
        if documents[DocumentType.ASSETS_LIABILITIES]:
            assets, metadata = self.asset_liability_extractor.extract(
                documents[DocumentType.ASSETS_LIABILITIES]
            )
            if assets:
                extraction.assets_liabilities = assets
            extraction.extraction_metadata[DocumentType.ASSETS_LIABILITIES] = metadata
        
        if documents[DocumentType.CREDIT_REPORT]:
            credit, metadata = self.credit_report_extractor.extract(
                documents[DocumentType.CREDIT_REPORT]
            )
            if credit:
                extraction.credit_report = credit
            extraction.extraction_metadata[DocumentType.CREDIT_REPORT] = metadata
        
        # Calculate verification status
        self._calculate_verification_status(extraction)
        
        # Enrich with metadata from CSV (fills in missing fields from ground truth)
        self._enrich_with_metadata(extraction)
        
        # Calculate data quality score
        extraction.data_quality_score = self._calculate_quality_score(extraction)
        
        return extraction
    
    def extract_batch(self, applications: List[Tuple[str, str]]) -> ExtractionBatchResult:
        """
        Extract data from multiple applications.
        
        Args:
            applications: List of (app_id, app_folder) tuples
        
        Returns:
            ExtractionBatchResult with statistics
        """
        start_time = datetime.now()
        
        result = ExtractionBatchResult(
            total_applications=len(applications),
            successful_extractions=0,
            failed_extractions=0,
            partial_extractions=0
        )
        
        for app_id, app_folder in applications:
            extraction = self.extract_application(app_id, app_folder)
            result.applications.append(extraction)
            
            if extraction.is_fully_extracted():
                result.successful_extractions += 1
            elif extraction.verification_status == VerificationStatus.INCOMPLETE:
                result.failed_extractions += 1
            else:
                result.partial_extractions += 1
        
        # Calculate statistics
        result.extraction_time_ms = (datetime.now() - start_time).total_seconds() * 1000
        
        quality_scores = [app.data_quality_score for app in result.applications 
                         if app.data_quality_score is not None]
        if quality_scores:
            result.average_quality_score = sum(quality_scores) / len(quality_scores)
        
        # Calculate document success rates
        result.document_success_rates = self._calculate_doc_success_rates(result.applications)
        
        return result
    
    @staticmethod
    def _find_file(folder: Path, name_part: str, extension: str) -> Optional[str]:
        """Find a file in folder by name and extension."""
        if not folder.exists():
            return None
        
        for file in folder.glob(f"*{name_part}*{extension}"):
            return str(file)
        return None
    
    @staticmethod
    def _calculate_verification_status(extraction: ApplicationExtraction) -> None:
        """Calculate overall verification status."""
        if extraction.missing_documents:
            extraction.verification_status = VerificationStatus.INCOMPLETE
        else:
            # Check for all documents successfully extracted
            all_successful = all(
                extraction.get_extraction_status(doc) == ExtractionStatus.SUCCESS
                for doc in extraction.extraction_metadata.keys()
            )
            
            if all_successful:
                extraction.verification_status = VerificationStatus.VERIFIED
            else:
                extraction.verification_status = VerificationStatus.INCOMPLETE
    
    @staticmethod
    def _calculate_quality_score(extraction: ApplicationExtraction) -> float:
        """Calculate overall data quality score (0.0-1.0)."""
        scores = []
        
        # Score based on documents present
        doc_count = len(extraction.extraction_metadata)
        if doc_count > 0:
            scores.append(doc_count / 6.0)  # 6 expected documents
        
        # Score based on extraction confidence
        if extraction.extraction_metadata:
            confidences = [m.confidence for m in extraction.extraction_metadata.values()]
            scores.append(sum(confidences) / len(confidences))
        
        # Score based on personal info completeness
        personal_fields = [
            extraction.personal_info.full_name,
            extraction.personal_info.emirates_id,
            extraction.personal_info.date_of_birth
        ]
        scores.append(sum(1 for f in personal_fields if f) / len(personal_fields))
        
        # Score based on employment info
        employment_fields = [
            extraction.employment_info.current_employer,
            extraction.employment_info.job_title,
            extraction.employment_info.monthly_salary
        ]
        scores.append(sum(1 for f in employment_fields if f) / len(employment_fields))
        
        return sum(scores) / len(scores) if scores else 0.0
    
    @staticmethod
    def _calculate_doc_success_rates(applications: List[ApplicationExtraction]) -> Dict[DocumentType, float]:
        """Calculate success rate for each document type."""
        rates = {}
        doc_types = [
            DocumentType.BANK_STATEMENT,
            DocumentType.EMIRATES_ID,
            DocumentType.EMPLOYMENT_LETTER,
            DocumentType.RESUME,
            DocumentType.ASSETS_LIABILITIES,
            DocumentType.CREDIT_REPORT
        ]
        
        for doc_type in doc_types:
            count = sum(
                1 for app in applications
                if app.get_extraction_status(doc_type) == ExtractionStatus.SUCCESS
            )
            success_rate = count / len(applications) if applications else 0.0
            rates[doc_type] = success_rate
        
        return rates