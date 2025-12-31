#!/usr/bin/env python3
"""
Build NetworkX graph with real nodes and edges from application data.
Structure: Person -> Application -> Documents -> Decision -> Recommendation
"""
import json
import sys
from pathlib import Path
from datetime import datetime, timedelta
import random

import networkx as nx
from openpyxl import load_workbook

sys.path.append(str(Path(__file__).parent))


def load_application_data(app_dir: Path) -> dict:
    """Load financial data from application directory."""
    
    # Load credit report for person data
    credit_path = app_dir / "credit_report.json"
    with open(credit_path, 'r') as f:
        credit_data = json.load(f)
    
    # Load assets/liabilities from Excel
    assets_path = app_dir / "assets_liabilities.xlsx"
    wb = load_workbook(assets_path, data_only=True)
    ws = wb.active
    
    monthly_income = 0
    total_assets = 0
    total_liabilities = 0
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0] or not row[1]:
            continue
        label = str(row[0]).lower()
        try:
            value = float(row[1])
        except (ValueError, TypeError):
            continue
        
        if "monthly income" in label:
            monthly_income = value
        elif "total assets" in label:
            total_assets = value
        elif "total liabilities" in label:
            total_liabilities = value
    
    # Extract person data
    subject = credit_data.get("subject", {})
    person_data = {
        'name': subject.get('name', 'Unknown'),
        'emirates_id': subject.get('emirates_id', ''),
        'nationality': subject.get('nationality', 'Unknown'),
        'monthly_income': monthly_income,
        'total_assets': total_assets,
        'total_liabilities': total_liabilities,
        'net_worth': total_assets - total_liabilities,
        'credit_score': credit_data.get('credit_score', 650)
    }
    
    return person_data


def calculate_policy_score(data: dict) -> float:
    """Calculate policy score for decision."""
    score = 0.0
    
    # Income (30 pts)
    if data['monthly_income'] < 3000:
        score += 10
    elif data['monthly_income'] < 10000:
        score += 20
    else:
        score += 30
    
    # Employment (20 pts) - estimate from income
    if data['monthly_income'] > 15000:
        score += 10  # Likely private high earner
    elif data['monthly_income'] > 8000:
        score += 15  # Likely government
    elif data['monthly_income'] > 3000:
        score += 5   # Self-employed
    
    # Family size (20 pts) - estimate
    expense_ratio = 0.6  # default
    family_size = 3
    if expense_ratio > 0.7:
        family_size = 5
        score += 20
    elif expense_ratio > 0.5:
        family_size = 3
        score += 10
    else:
        family_size = 1
        score += 5
    
    # Net worth (15 pts)
    if data['net_worth'] < 50000:
        score += 15
    elif data['net_worth'] < 150000:
        score += 10
    else:
        score += 5
    
    # Credit (15 pts)
    if data['credit_score'] < 550:
        score += 5
    elif data['credit_score'] < 700:
        score += 10
    else:
        score += 15
    
    return min(score, 100)


def build_graph(docs_path: Path, output_path: Path):
    """Build NetworkX graph from application data."""
    
    print("\n" + "="*80)
    print("BUILDING NETWORKX GRAPH - PRODUCTION DATA MODEL")
    print("="*80)
    
    # Create directed graph
    G = nx.DiGraph()
    
    # Metadata
    G.graph['name'] = 'Social Support Application Network'
    G.graph['created_at'] = datetime.now().isoformat()
    G.graph['description'] = 'Graph representation of social support applications with person, document, decision nodes'
    
    print("\n1. Processing applications and creating nodes...")
    
    app_dirs = sorted([d for d in docs_path.iterdir() if d.is_dir() and d.name.startswith("APP-")])
    
    stats = {
        'person_nodes': 0,
        'application_nodes': 0,
        'document_nodes': 0,
        'decision_nodes': 0,
        'recommendation_nodes': 0,
        'edges': 0
    }
    
    for app_dir in app_dirs:
        app_id = app_dir.name
        
        # Load data
        person_data = load_application_data(app_dir)
        policy_score = calculate_policy_score(person_data)
        
        # Determine decision
        if policy_score < 30:
            decision = "DECLINED"
            priority = "LOW"
        elif policy_score < 50:
            decision = "CONDITIONAL"
            priority = "MEDIUM"
        elif policy_score < 70:
            decision = "APPROVED"
            priority = "MEDIUM"
        else:
            decision = "APPROVED"
            priority = "HIGH"
        
        # Create person node
        person_id = person_data['emirates_id']
        G.add_node(
            person_id,
            node_type='Person',
            name=person_data['name'],
            nationality=person_data['nationality'],
            monthly_income=person_data['monthly_income'],
            credit_score=person_data['credit_score'],
            net_worth=person_data['net_worth']
        )
        stats['person_nodes'] += 1
        
        # Create application node
        submission_date = datetime.now() - timedelta(days=random.randint(1, 90))
        G.add_node(
            app_id,
            node_type='Application',
            submitted_by=person_id,
            submission_date=submission_date.isoformat(),
            status='PROCESSED',
            policy_score=policy_score
        )
        stats['application_nodes'] += 1
        
        # Edge: Person -> Application (SUBMITTED)
        G.add_edge(person_id, app_id, relationship='SUBMITTED', timestamp=submission_date.isoformat())
        stats['edges'] += 1
        
        # Create document nodes
        doc_types = ['bank_statement', 'emirates_id', 'employment_letter', 'resume', 'assets_liabilities', 'credit_report']
        for doc_type in doc_types:
            doc_id = f"{app_id}_{doc_type}"
            
            # Find actual file
            doc_files = list(app_dir.glob(f"{doc_type}.*"))
            doc_file = doc_files[0].name if doc_files else f"{doc_type}.pdf"
            
            G.add_node(
                doc_id,
                node_type='Document',
                document_type=doc_type,
                filename=doc_file,
                file_path=str(app_dir / doc_file),
                uploaded_at=submission_date.isoformat()
            )
            stats['document_nodes'] += 1
            
            # Edge: Application -> Document (HAS_DOCUMENT)
            G.add_edge(app_id, doc_id, relationship='HAS_DOCUMENT')
            stats['edges'] += 1
        
        # Create decision node
        decision_id = f"DECISION_{app_id}"
        review_date = submission_date + timedelta(days=random.randint(1, 7))
        G.add_node(
            decision_id,
            node_type='Decision',
            application_id=app_id,
            decision=decision,
            policy_score=policy_score,
            decided_at=review_date.isoformat(),
            decided_by='SYSTEM',
            priority=priority
        )
        stats['decision_nodes'] += 1
        
        # Edge: Application -> Decision (REVIEWED)
        G.add_edge(app_id, decision_id, relationship='REVIEWED', timestamp=review_date.isoformat())
        stats['edges'] += 1
        
        # Create recommendation node if approved/conditional
        if decision in ['APPROVED', 'CONDITIONAL']:
            rec_id = f"REC_{app_id}"
            
            # Calculate support amount
            if decision == 'APPROVED':
                support_amount = min(person_data['monthly_income'] * 0.3, 3000)
            else:
                support_amount = min(person_data['monthly_income'] * 0.2, 2000)
            
            G.add_node(
                rec_id,
                node_type='Recommendation',
                application_id=app_id,
                support_type='FINANCIAL_ASSISTANCE',
                support_amount=round(support_amount, 2),
                duration_months=6 if decision == 'APPROVED' else 3,
                created_at=review_date.isoformat()
            )
            stats['recommendation_nodes'] += 1
            
            # Edge: Decision -> Recommendation (RECOMMENDED)
            G.add_edge(decision_id, rec_id, relationship='RECOMMENDED')
            stats['edges'] += 1
    
    print(f"   ✓ Created {stats['person_nodes']} person nodes")
    print(f"   ✓ Created {stats['application_nodes']} application nodes")
    print(f"   ✓ Created {stats['document_nodes']} document nodes")
    print(f"   ✓ Created {stats['decision_nodes']} decision nodes")
    print(f"   ✓ Created {stats['recommendation_nodes']} recommendation nodes")
    print(f"   ✓ Created {stats['edges']} edges")
    
    # Graph metrics
    print("\n2. Graph Metrics:")
    print(f"   Nodes: {G.number_of_nodes()}")
    print(f"   Edges: {G.number_of_edges()}")
    print(f"   Density: {nx.density(G):.4f}")
    print(f"   Avg Degree: {sum(dict(G.degree()).values()) / G.number_of_nodes():.2f}")
    
    # Save graph
    print("\n3. Saving graph data...")
    
    # GraphML (for visualization tools)
    graphml_path = output_path / "application_graph.graphml"
    nx.write_graphml(G, graphml_path)
    print(f"   ✓ GraphML saved: {graphml_path}")
    
    # JSON (for FastAPI/Streamlit)
    json_data = nx.node_link_data(G)
    json_path = output_path / "application_graph.json"
    with open(json_path, 'w') as f:
        json.dump(json_data, f, indent=2)
    print(f"   ✓ JSON saved: {json_path}")
    
    # Statistics JSON
    stats_data = {
        'graph_stats': {
            'total_nodes': G.number_of_nodes(),
            'total_edges': G.number_of_edges(),
            'density': nx.density(G),
            'avg_degree': sum(dict(G.degree()).values()) / G.number_of_nodes()
        },
        'node_counts': stats,
        'node_types': {
            'Person': stats['person_nodes'],
            'Application': stats['application_nodes'],
            'Document': stats['document_nodes'],
            'Decision': stats['decision_nodes'],
            'Recommendation': stats['recommendation_nodes']
        },
        'relationship_types': {
            'SUBMITTED': stats['application_nodes'],
            'HAS_DOCUMENT': stats['document_nodes'],
            'REVIEWED': stats['decision_nodes'],
            'RECOMMENDED': stats['recommendation_nodes']
        }
    }
    
    stats_path = output_path / "graph_statistics.json"
    with open(stats_path, 'w') as f:
        json.dump(stats_data, f, indent=2)
    print(f"   ✓ Statistics saved: {stats_path}")
    
    # Sample queries
    print("\n4. Sample Graph Queries:")
    
    # Find all approved applications
    approved = [n for n, d in G.nodes(data=True) 
                if d.get('node_type') == 'Decision' and d.get('decision') == 'APPROVED']
    print(f"   Approved applications: {len(approved)}")
    
    # Find high-priority cases
    high_priority = [n for n, d in G.nodes(data=True) 
                     if d.get('node_type') == 'Decision' and d.get('priority') == 'HIGH']
    print(f"   High-priority cases: {len(high_priority)}")
    
    # Average policy score
    scores = [d['policy_score'] for n, d in G.nodes(data=True) 
              if d.get('node_type') == 'Application']
    avg_score = sum(scores) / len(scores) if scores else 0
    print(f"   Average policy score: {avg_score:.1f}")
    
    print("\n" + "="*80)
    print("✓ NETWORKX GRAPH BUILD COMPLETE")
    print("="*80)
    print(f"\nGraph accessible via:")
    print(f"  - NetworkX: Load from {graphml_path}")
    print(f"  - FastAPI: Read from {json_path}")
    print(f"  - Streamlit: Use nx.node_link_graph(json.load('{json_path}'))")
    print()
    
    return G


if __name__ == "__main__":
    docs_path = Path(__file__).parent / "data" / "processed" / "documents"
    output_path = Path(__file__).parent / "data" / "databases"
    output_path.mkdir(exist_ok=True, parents=True)
    
    if not docs_path.exists() or not any(docs_path.iterdir()):
        print("ERROR: No application data found. Run generate_stratified_dataset.py first.")
        sys.exit(1)
    
    build_graph(docs_path, output_path)
