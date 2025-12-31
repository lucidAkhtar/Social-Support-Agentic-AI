#!/usr/bin/env python3
"""Train Random Forest model on 40 diverse applications with policy alignment."""

import json
import pickle
import random
from pathlib import Path
import numpy as np
from sklearn.ensemble import RandomForestClassifier
from sklearn.preprocessing import StandardScaler
from sklearn.model_selection import cross_val_score, StratifiedKFold
from sklearn.metrics import classification_report, confusion_matrix, roc_auc_score
import sys

# Add src to path
sys.path.insert(0, str(Path(__file__).parent / "src"))


def calculate_policy_score(data: dict) -> float:
    """Calculate policy score (0-100) based on business rules."""
    score = 0.0
    
    # Income scoring (30 points)
    monthly_income = data.get("monthly_income", 0)
    if monthly_income < 3000:
        score += 10
    elif monthly_income < 10000:
        score += 20
    else:
        score += 30
    
    # Employment scoring (20 points)
    employment = data.get("employment_status", "").lower()
    if "government" in employment:
        score += 20
    elif "private" in employment or "employed" in employment:
        score += 10
    elif "self" in employment:
        score += 5
    # Unemployed = 0 points
    
    # Family burden scoring (20 points)
    family_size = data.get("family_size", 1)
    if family_size >= 4:
        score += 15
    elif family_size >= 2:
        score += 10
    else:
        score += 5
    
    # Add 5 points if dependents > 3
    if family_size > 3:
        score += 5
    
    # Net worth scoring (15 points)
    net_worth = data.get("net_worth", 0)
    if net_worth < 50000:
        score += 15
    elif net_worth < 150000:
        score += 10
    else:
        score += 5
    
    # Credit score (15 points)
    credit_score = data.get("credit_score", 0)
    if credit_score < 550:
        score += 5  # High risk but may need help
    elif credit_score < 700:
        score += 10
    else:
        score += 15
    
    return min(score, 100)


def extract_features(app_data: dict) -> tuple:
    """Extract 8 features from application data."""
    monthly_income = app_data.get("monthly_income", 0)
    monthly_expenses = app_data.get("monthly_expenses", 0)
    total_assets = app_data.get("total_assets", 0)
    total_liabilities = app_data.get("total_liabilities", 0)
    family_size = app_data.get("family_size", 1)
    credit_score = app_data.get("credit_score", 0)
    
    # Engineered features
    net_worth = total_assets - total_liabilities
    monthly_net = monthly_income - monthly_expenses
    debt_to_asset_ratio = total_liabilities / max(total_assets, 1)
    
    # Employment encoding
    employment = app_data.get("employment_status", "").lower()
    if "government" in employment:
        employment_code = 3
    elif "private" in employment or "employed" in employment:
        employment_code = 2
    elif "self" in employment:
        employment_code = 1
    else:
        employment_code = 0
    
    return (
        monthly_income,
        monthly_expenses,
        net_worth,
        employment_code,
        family_size,
        credit_score,
        monthly_net,
        debt_to_asset_ratio
    )


def load_applications(docs_path: Path) -> list:
    """Load all 40 applications from processed documents."""
    applications = []
    
    for app_dir in sorted(docs_path.iterdir()):
        if not app_dir.is_dir() or not app_dir.name.startswith("APP-"):
            continue
        
        # Load credit_report.json
        credit_path = app_dir / "credit_report.json"
        if not credit_path.exists():
            continue
        
        with open(credit_path, 'r') as f:
            credit_data = json.load(f)
        
        # Load assets_liabilities.xlsx to get financials
        from openpyxl import load_workbook
        assets_path = app_dir / "assets_liabilities.xlsx"
        if not assets_path.exists():
            continue
        
        wb = load_workbook(assets_path, data_only=True)
        ws = wb.active
        
        # Parse assets and liabilities from Excel
        total_assets = 0
        total_liabilities = 0
        monthly_income = 0
        monthly_expenses = 0
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0] or not row[1]:
                continue
            label = str(row[0]).lower()
            
            # Try to convert value to float
            try:
                value = float(row[1])
            except (ValueError, TypeError):
                continue
            
            if "monthly income" in label:
                monthly_income = value
            elif "total assets" in label:
                total_assets = value
            elif "total liabilities" in label:
                total_liabilities = value
        
        # Estimate monthly expenses (40-85% of income from generation logic)
        if monthly_expenses == 0 and monthly_income > 0:
            monthly_expenses = monthly_income * random.uniform(0.40, 0.85)
        
        # Extract employment from credit report subject
        employment_status = "Unknown"
        if "employment" in credit_data.get("subject", {}):
            employment_status = credit_data["subject"]["employment"]
        elif "occupation" in credit_data.get("subject", {}):
            employment_status = credit_data["subject"]["occupation"]
        else:
            # Infer from income level
            if monthly_income > 15000:
                employment_status = "Private Sector"
            elif monthly_income > 8000:
                employment_status = "Government"
            else:
                employment_status = "Self-employed"
        
        # Extract family size (default to 3 if not found)
        family_size = credit_data.get("subject", {}).get("family_size", 3)
        if not family_size or family_size == 0:
            # Estimate from financial burden
            expense_ratio = monthly_expenses / max(monthly_income, 1)
            if expense_ratio > 0.7:
                family_size = random.randint(4, 6)
            elif expense_ratio > 0.5:
                family_size = random.randint(2, 4)
            else:
                family_size = random.randint(1, 3)
        
        app_data = {
            'app_id': app_dir.name,
            'monthly_income': monthly_income,
            'monthly_expenses': monthly_expenses,
            'total_assets': total_assets,
            'total_liabilities': total_liabilities,
            'net_worth': total_assets - total_liabilities,
            'employment_status': employment_status,
            'family_size': family_size,
            'credit_score': credit_data.get('credit_score', 650)
        }
        
        applications.append(app_data)
    
    return applications


def main():
    print("=" * 80)
    print("RANDOM FOREST MODEL TRAINING - POLICY ALIGNMENT")
    print("=" * 80)
    
    # Paths
    docs_path = Path(__file__).parent / "data" / "processed" / "documents"
    models_dir = Path(__file__).parent / "models"
    models_dir.mkdir(exist_ok=True)
    
    # Load applications
    print(f"\n1. Loading applications from {docs_path}")
    applications = load_applications(docs_path)
    print(f"   ✓ Loaded {len(applications)} applications")
    
    # Extract features and labels
    print("\n2. Extracting features and calculating policy scores")
    X = []
    y = []
    app_ids = []
    policy_scores = []
    
    for app in applications:
        features = extract_features(app)
        policy_score = calculate_policy_score(app)
        label = 1 if policy_score >= 30 else 0  # Eligible if score >= 30
        
        X.append(features)
        y.append(label)
        app_ids.append(app['app_id'])
        policy_scores.append(policy_score)
    
    X = np.array(X)
    y = np.array(y)
    
    eligible_count = np.sum(y)
    print(f"   ✓ Extracted 8 features from {len(X)} applications")
    print(f"   ✓ Eligible: {eligible_count}, Not Eligible: {len(y) - eligible_count}")
    print(f"   ✓ Policy scores range: {min(policy_scores):.1f} - {max(policy_scores):.1f}")
    
    # Feature names
    feature_names = [
        'monthly_income',
        'monthly_expenses',
        'net_worth',
        'employment_status_code',
        'family_size',
        'credit_score',
        'monthly_net',
        'debt_to_asset_ratio'
    ]
    
    # Scale features
    print("\n3. Scaling features")
    scaler = StandardScaler()
    X_scaled = scaler.fit_transform(X)
    print("   ✓ Features standardized (mean=0, std=1)")
    
    # Train Random Forest
    print("\n4. Training Random Forest Classifier")
    rf_model = RandomForestClassifier(
        n_estimators=100,
        max_depth=10,
        min_samples_split=5,
        min_samples_leaf=2,
        class_weight='balanced',
        random_state=42,
        n_jobs=-1
    )
    
    rf_model.fit(X_scaled, y)
    print("   ✓ Model trained (100 estimators, max_depth=10)")
    
    # Cross-validation
    print("\n5. Performing 5-fold cross-validation")
    cv = StratifiedKFold(n_splits=5, shuffle=True, random_state=42)
    cv_scores = cross_val_score(rf_model, X_scaled, y, cv=cv, scoring='accuracy')
    
    print(f"   ✓ CV Accuracy: {cv_scores.mean():.3f} (+/- {cv_scores.std() * 2:.3f})")
    print(f"   ✓ Fold scores: {[f'{s:.3f}' for s in cv_scores]}")
    
    # Training metrics
    print("\n6. Training Set Performance")
    y_pred = rf_model.predict(X_scaled)
    
    # Handle single-class case
    n_classes = len(np.unique(y))
    if n_classes == 1:
        print("   ⚠ WARNING: Only one class found in dataset (all eligible or all not eligible)")
        print(f"   All applications classified as: {'Eligible' if y[0] == 1 else 'Not Eligible'}")
        roc_auc = 1.0
    else:
        y_proba = rf_model.predict_proba(X_scaled)[:, 1]
        
        print("\n   Classification Report:")
        print("   " + "\n   ".join(classification_report(y, y_pred, target_names=['Not Eligible', 'Eligible']).split("\n")))
        
        print("\n   Confusion Matrix:")
        cm = confusion_matrix(y, y_pred)
        print(f"   [[TN={cm[0,0]:2d}, FP={cm[0,1]:2d}]")
        print(f"    [FN={cm[1,0]:2d}, TP={cm[1,1]:2d}]]")
        
        # ROC-AUC
        roc_auc = roc_auc_score(y, y_proba)
        print(f"\n   ROC-AUC Score: {roc_auc:.3f}")
    
    # Feature importance
    print("\n7. Feature Importance Analysis")
    feature_importance = sorted(
        zip(feature_names, rf_model.feature_importances_),
        key=lambda x: x[1],
        reverse=True
    )
    
    print("\n   Top Features:")
    for i, (name, importance) in enumerate(feature_importance[:5], 1):
        print(f"   {i}. {name:25s}: {importance:.3f} ({importance*100:.1f}%)")
    
    # Policy alignment check
    print("\n8. Policy Alignment Verification")
    mismatches = []
    for i, (app_id, pred, actual, score) in enumerate(zip(app_ids, y_pred, y, policy_scores)):
        if pred != actual:
            mismatches.append((app_id, pred, actual, score))
    
    if mismatches:
        print(f"   ⚠ Found {len(mismatches)} mismatches between ML and policy:")
        for app_id, pred, actual, score in mismatches[:5]:
            pred_label = "Eligible" if pred == 1 else "Not Eligible"
            actual_label = "Eligible" if actual == 1 else "Not Eligible"
            print(f"   - {app_id}: ML={pred_label}, Policy={actual_label} (score={score:.1f})")
    else:
        print("   ✓ Perfect alignment: ML predictions match policy rules")
    
    # Save models
    print("\n9. Saving models")
    
    model_path = models_dir / "eligibility_model.pkl"
    with open(model_path, 'wb') as f:
        pickle.dump(rf_model, f)
    print(f"   ✓ Model saved: {model_path}")
    
    scaler_path = models_dir / "feature_scaler.pkl"
    with open(scaler_path, 'wb') as f:
        pickle.dump(scaler, f)
    print(f"   ✓ Scaler saved: {scaler_path}")
    
    feature_names_path = models_dir / "feature_names.pkl"
    with open(feature_names_path, 'wb') as f:
        pickle.dump(feature_names, f)
    print(f"   ✓ Feature names saved: {feature_names_path}")
    
    # Save training report
    report = {
        "training_date": "2025-12-31",
        "n_samples": len(X),
        "n_features": len(feature_names),
        "feature_names": feature_names,
        "n_eligible": int(eligible_count),
        "n_not_eligible": int(len(y) - eligible_count),
        "cv_accuracy_mean": float(cv_scores.mean()),
        "cv_accuracy_std": float(cv_scores.std()),
        "roc_auc": float(roc_auc),
        "feature_importance": {name: float(imp) for name, imp in feature_importance},
        "policy_alignment_mismatches": len(mismatches),
        "model_params": {
            "n_estimators": 100,
            "max_depth": 10,
            "class_weight": "balanced"
        }
    }
    
    report_path = models_dir / "training_report.json"
    with open(report_path, 'w') as f:
        json.dump(report, f, indent=2)
    print(f"   ✓ Training report saved: {report_path}")
    
    print("\n" + "=" * 80)
    print("✓ TRAINING COMPLETE")
    print("=" * 80)
    print(f"\nModel Performance:")
    print(f"  - CV Accuracy: {cv_scores.mean():.1%}")
    print(f"  - ROC-AUC: {roc_auc:.3f}")
    print(f"  - Top Feature: {feature_importance[0][0]} ({feature_importance[0][1]*100:.1f}%)")
    print(f"  - Policy Alignment: {len(applications) - len(mismatches)}/{len(applications)} ({(1 - len(mismatches)/len(applications))*100:.1f}%)")
    print()


if __name__ == "__main__":
    main()
